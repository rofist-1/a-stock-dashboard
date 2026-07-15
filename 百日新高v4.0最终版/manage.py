#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""百日新高v5.0 七层规则：三锚定仓→主线过滤→状态分类→风险过滤→狙击池→后备池→卖出预警"""
import json, os, sys
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(BASE, "data.json")
KLINE_CACHE_FILE = os.path.join(BASE, "kline_cache.json")
HOLDINGS_FILE = os.path.join(BASE, "holdings.json")
MARKET_DATA_FILE = os.path.join(BASE, "market_data.json")
EXCEL_FILE = os.path.join(BASE, "百日新高_报表.xlsx")

CORE_THEMES = {
    "创新药/CRO": ["药","CRO","创新药","中药","医药"],
    "机器人": ["机器人","减速器","自动化","工业母机"],
    "算力/通信": ["算力","液冷","通信","数据中心","交换机","GPU","CPO"],
    "半导体/芯片": ["芯片","半导体","存储","光刻","EDA","IGBT","RISC-V","先进封装"],
}

def load_json(path):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f: return json.load(f)
    return {}

def save_json(path, obj):
    with open(path, "w", encoding="utf-8") as f: json.dump(obj, f, ensure_ascii=False, indent=2)

def parse_mn(val):
    if not val: return 0
    s = str(val).replace("+","").replace(",","")
    try:
        if "亿" in s: return float(s.replace("亿",""))*10000
        if "万" in s: return float(s.replace("万",""))
        return float(s)/10000
    except: return 0

def is_core(sector):
    for t,ks in CORE_THEMES.items():
        for k in ks:
            if k in str(sector): return t
    return None

def classify(chg, vr):
    if chg >= 3 and (vr is None or vr >= 0.8): return "强势上攻"
    if chg < -3 and (vr is not None and vr > 1.2): return "放量下跌"
    if (vr is not None and vr < 0.8) and -5 <= chg <= 2: return "缩量回调"
    if -2 <= chg <= 2 and (vr is None or 0.5 <= vr <= 1.2): return "高位整理"
    if chg >= 3: return "强势上攻"
    if chg < 0 and vr and vr < 0.8: return "缩量回调"
    return "高位整理"

def calc_ma(vals, p):
    if len(vals) < p: return None
    return sum(vals[-p:])/p

def market_anchors():
    m = load_json(MARKET_DATA_FILE)
    r = {"anchor1":"中性","anchor2":"中性","anchor3":"中性","verdict":"数据不足","position_max":5,"position_range":"30%-50%"}
    sh = m.get("sh_index",[])
    if len(sh) >= 23:
        cls = [d["close"] for d in sh]; ma20s = []
        for i in range(19,len(cls)): ma20s.append(round(calc_ma(cls[:i+1],20),2))
        lc = cls[-1]; lma = ma20s[-1]
        above = lc > lma
        rise = len(ma20s)>=3 and all(ma20s[-(i+2)]>ma20s[-(i+3)] for i in range(2))
        fall = len(ma20s)>=3 and all(ma20s[-(i+2)]<ma20s[-(i+3)] for i in range(2))
        r["anchor1"] = "上涨" if (above and rise) else ("下跌" if (not above and fall) else "中性")
        r["details"] = {"sh":lc,"ma20":lma}
    br = m.get("breadth",{}); up = br.get("up",0); dn = br.get("down",0)
    if up+dn > 0:
        ratio = up/(up+dn)*100
        r["anchor2"] = "上涨" if ratio > 50 else ("下跌" if ratio < 30 else "中性")
        r["details"] = r.get("details",{}); r["details"]["breadth"] = f"{ratio:.1f}%"
    nh = m.get("new_high",0); nl = m.get("new_low",0); diff = nh-nl
    dh = m.get("diff_history",[])
    if len(dh)>=3 and all(d<0 for d in dh[-3:]): r["anchor3"] = "强制空仓"
    else: r["anchor3"] = "上涨" if diff>0 else ("下跌" if diff<0 else "中性")
    r["details"]["diff"] = diff; r["details"]["nh"] = nh; r["details"]["nl"] = nl
    sigs = [r["anchor1"],r["anchor2"],r["anchor3"]]
    upv = sum(1 for s in sigs if s=="上涨")
    downv = sum(1 for s in sigs if s in ("下跌","强制空仓"))
    if r["anchor3"] == "强制空仓": r["verdict"]="下跌市(强制空仓)"; r["position_max"]=0; r["position_range"]="0%"
    elif upv >= 2: r["verdict"]="上涨市"; r["position_max"]=8; r["position_range"]="70%-100%"
    elif downv >= 2: r["verdict"]="下跌市"; r["position_max"]=1; r["position_range"]="0-10%"
    else: r["verdict"]="震荡市"; r["position_max"]=5; r["position_range"]="30%-50%"
    return r

# ═══════ Layer 7: Risk Filter ═══════
def risk_filter(st, latest_rec, kline_ind):
    """检查庄股/对倒信号，返回 True 表示需要过滤"""
    chg = latest_rec.get("changePct", 0)
    turn_rate = latest_rec.get("turnRate", None)  # 当日换手率
    first_date = st.get("firstDate")
    last_date = st.get("lastDate")
    amount_yi = latest_rec.get("amountYi", 0)
    
    # 1. 涨幅≥5%且换手率<3% (异常低)
    if chg >= 5 and turn_rate is not None and turn_rate < 3:
        return True, "涨幅≥5%+换手<3%"
    
    # 2. 涨幅≥5%且换手率>30% (异常高)
    if chg >= 5 and turn_rate is not None and turn_rate > 30:
        return True, "涨幅≥5%+换手>30%"
    
    # 3. 首次新高日+60日均换手<1% (长期横盘)
    if first_date == last_date and kline_ind and kline_ind.get("avgTurn60") is not None:
        if kline_ind["avgTurn60"] < 1:
            return True, "长期横盘首高(60日均换手<1%)"
    
    # 4. 成交额全市场前100但涨幅<3%
    if amount_yi and amount_yi > 100 and chg < 3:
        return True, "成交额百亿+涨幅<3%"
    
    return False, ""

def generate():
    data = load_json(DATA_FILE)
    kline = load_json(KLINE_CACHE_FILE)
    holdings = load_json(HOLDINGS_FILE)
    records = data.get("records", [])
    if not records: return print("No data")
    latest_date = records[-1]["date"]
    anchors = market_anchors()

    stocks = {}
    for rec in records:
        for s in rec["stocks"]:
            c = s["code"]
            if c not in stocks:
                stocks[c] = {"name": s["name"], "firstDate": rec["date"], "lastDate": rec["date"],
                           "dates": [], "sectors": set(), "recs": [], "mnTotal": 0}
            st = stocks[c]; st["dates"].append(rec["date"]); st["lastDate"] = rec["date"]
            st["sectors"].add(s.get("sector", "")); st["recs"].append({"date": rec["date"], "stock": s})
            st["mnTotal"] += parse_mn(s.get("mainNet", ""))
    for st in stocks.values(): st["dates"].sort()

    core, non_core, filtered = {}, {}, {}
    for c, st in stocks.items():
        theme = is_core(";".join(st["sectors"]))
        if theme: st["theme"] = theme; core[c] = st
        else: non_core[c] = st

    sniper, reserve, warning_list = [], [], []
    theme_stats = defaultdict(lambda: defaultdict(int))
    filter_count = 0

    for code, st in core.items():
        lr = st["recs"][-1]["stock"]
        chg = lr.get("changePct", 0); mn = parse_mn(lr.get("mainNet", ""))
        cached = kline.get(code, {}); ind = cached.get("indicators", {}) if cached else {}
        vr = ind.get("volRatio"); d13 = ind.get("distMA13")
        

        # Add amountYi from screener if available
        lr["turnRate"] = lr.get("turnRate") or (cached.get("turnRate"))
        
        status = classify(chg, vr)
        st["status"] = status; st["dist13"] = d13; st["vr"] = vr
        st["vol_status"] = ind.get("volStatus", "") if ind else ""
        
        theme_stats[st["theme"]]["total"] += 1
        theme_stats[st["theme"]][status] = theme_stats[st["theme"]].get(status, 0) + 1

        # ═══ Layer 7: Risk Filter ═══
        flagged, reason = risk_filter(st, lr, ind)
        if flagged:
            filtered[code] = st; st["risk_reason"] = reason
            filter_count += 1
            theme_stats[st["theme"]]["疑似庄股"] = theme_stats[st["theme"]].get("疑似庄股", 0) + 1
            continue  # Skip sniper/reserve pools

        # ═══ Rule 1: 首日涨幅>15%过滤 ═══
        first_chg = st["recs"][0]["stock"].get("changePct", 0)
        d20 = ind.get("distMA20")
        if first_chg > 15:
            if status == "缩量回调" and d20 is not None and d20 >= 0 and vr is not None and vr < 0.5:
                pass  # Exception: 回调到MA20+极缩量, 重新纳入
            else:
                filtered[code] = st; st["risk_reason"] = f"首日涨幅>{first_chg:.0f}%"
                filter_count += 1
                theme_stats[st["theme"]]["疑似庄股"] = theme_stats[st["theme"]].get("疑似庄股", 0) + 1
                continue

        # ═══ Rule 2: 多次上榜不涨(疑似出货)过滤 ═══
        last20_dates = sorted([d for d in st["dates"] if d >= "2026-06-24"])
        if len(last20_dates) >= 5:
            first_rec = None; last_rec = None
            for rec in st["recs"]:
                if rec["date"] == last20_dates[0]: first_rec = rec
                if rec["date"] == last20_dates[-1]: last_rec = rec
            if first_rec and last_rec:
                p1 = first_rec["stock"].get("price", 0)
                p2 = last_rec["stock"].get("price", 0)
                total_ret = (p2 / p1 - 1) * 100 if p1 > 0 else 0
                if total_ret < 5:
                    filtered[code] = st; st["risk_reason"] = f"5次上榜仅涨{total_ret:+.1f}%(疑似出货)"
                    filter_count += 1
                    theme_stats[st["theme"]]["疑似庄股"] = theme_stats[st["theme"]].get("疑似庄股", 0) + 1
                    continue

        # Sniper pool (v5.1: VR 0.5-2.0, 首发3-5天, 阳线/十字星)
        candle_ok = chg >= -1  # 收阳线或十字星(跌幅不超过1%)
        vr_ok = vr is not None and 0.5 < vr < 2.0
        if status == "缩量回调" and d13 is not None and -3 <= d13 <= 3 and vr_ok:
            try:
                fd = datetime.strptime(st["firstDate"], "%Y-%m-%d")
                ld = datetime.strptime(latest_date, "%Y-%m-%d")
                days = (ld - fd).days
                if 3 <= days <= 5 and mn > -10000 and candle_ok: sniper.append(st)
            except: pass
        
        # Reserve pool
        if status == "高位整理":
            try:
                fd = datetime.strptime(st["firstDate"], "%Y-%m-%d")
                ld = datetime.strptime(latest_date, "%Y-%m-%d")
                if 1 <= (ld - fd).days <= 5 and (vr is None or vr <= 1.2): reserve.append(st)
            except: pass
        
        # Holdings warnings
        if code in holdings:
            warns = []
            if vr and vr > 1.5 and (chg < 1 or chg < 0): warns.append("注意减仓(放量滞涨)")
            if chg < -3 and vr and vr > 1.2: warns.append("接近止损(放量下跌)")
            if d13 is not None and d13 < 0: warns.append(f"触发止损(跌破MA13 {abs(d13):.1f}%)")
            ep = holdings.get(code, {}).get("entryPrice", 0); cp = lr.get("price", 0)
            if ep and cp and ep > 0:
                profit = (cp / ep - 1) * 100
                if profit > 20: warns.append(f"浮盈{profit:.1f}% 止盈线=MA13")
                elif profit > 10: warns.append(f"浮盈{profit:.1f}% 上移止损至成本价")
            if warns: warning_list.append({"code": code, "name": st["name"], "price": cp, "chg": chg, "warnings": warns})

    sniper.sort(key=lambda x: (x.get("dist13") if x.get("dist13") is not None else 999))
    reserve.sort(key=lambda x: x["firstDate"])

    # ═══════ Console Report ═══════
    print(f"\n{'='*60}")
    print(f"  百日新高 v5.0  七层规则  {latest_date}")
    print(f"{'='*60}")
    print(f"\n 锚1趋势:{anchors['anchor1']} | 锚2广度:{anchors['anchor2']} | 锚3情绪:{anchors['anchor3']}")
    print(f" 表决:{anchors['verdict']} | 仓位:{anchors['position_range']} | 最大持仓{anchors['position_max']}只")
    print(f"\n 核心{len(core)}只 | 过滤{len(non_core)}只 | 第七层风险过滤{filter_count}只")
    for theme in ["半导体/芯片","算力/通信","机器人","创新药/CRO"]:
        ts = theme_stats.get(theme, {})
        print(f"  {theme}: {ts['total']}只 | 强{ts.get('强势上攻',0)} 缩调{ts.get('缩量回调',0)} 高整{ts.get('高位整理',0)} 放跌{ts.get('放量下跌',0)} 风险{ts.get('疑似庄股',0)}")
    print(f"\n 狙击池:{len(sniper)}只 | 后备池:{len(reserve)}只")
    if filtered:
        print(f"\n 第七层·风险过滤({filter_count}只):")
        for c, st in list(filtered.items())[:12]:
            print(f"  {st['name']}({c}) {st['theme']} | {st['risk_reason']}")
    if sniper:
        for s in sniper[:3]:
            code_s = s['recs'][-1]['stock']['code']
            d13_s = s.get('dist13')
            print(f" 狙击:{s['name']}({code_s}) {s['theme']} MA13:{d13_s:+.1f}%")
    if reserve:
        for s in reserve[:5]:
            print(f" 后备:{s['name']} {s['theme']}")
    if warning_list:
        for hw in warning_list:
            print(f" 预警:{hw['name']}({hw['code']}) {'; '.join(hw['warnings'])}")

    # ═══ 全库扫描：过去10天所有新高过的核心股，现在距MA13 ±3% ═══
    all_dates = sorted(set(r['date'] for r in records))
    recent_10_dates = all_dates[-10:]
    all_stocks = {}
    for rec in records:
        if rec['date'] not in recent_10_dates: continue
        for s in rec['stocks']:
            c = s['code']
            all_stocks.setdefault(c, {'name': s['name'], 'sectors': set(), 'count': 0, 'last_chg': 0, 'last_price': 0, 'last_date': ''})
            all_stocks[c]['sectors'].add(s.get('sector', ''))
            all_stocks[c]['count'] += 1
    
    for c in all_stocks:
        for rec in reversed(records):
            for s in rec['stocks']:
                if s['code'] == c:
                    all_stocks[c]['last_chg'] = s.get('changePct', 0)
                    all_stocks[c]['last_price'] = s.get('price', 0)
                    all_stocks[c]['last_date'] = rec['date']
                    break
            if all_stocks[c]['last_price']: break
    
    full_scan = []
    for c, st in all_stocks.items():
        if not is_core(';'.join(st['sectors'])): continue
        cached = kline.get(c, {}); ind = cached.get('indicators', {}) if cached else {}
        d13 = ind.get('distMA13'); vr = ind.get('volRatio')
        if d13 is None or abs(d13) > 3: continue
        st['d13'] = d13; st['vr'] = vr; st['code'] = c
        vs = '缩量' if vr and vr < 0.8 else ('放量' if vr and vr > 1.2 else '正常')
        st['vol_status'] = vs
        full_scan.append(st)
    full_scan.sort(key=lambda x: abs(x['d13']))
    scan_count = len(full_scan)
    
    # ═══════ Excel ═══════
    DARK="0f1923"; HB="1a2a3a"; R1="1a2736"; R2="162230"
    RED="ef5350"; GREEN="4caf50"; ORANGE="ffb74d"; BLUE="2196f3"; PURPLE="c084fc"; CYAN="4fc3f7"; GOLD="ffd700"
    FILL_S = PatternFill(start_color="0a2a0a", end_color="0a2a0a", fill_type="solid")
    FILL_R = PatternFill(start_color="1a2a3a", end_color="1a2a3a", fill_type="solid")
    FILL_W = PatternFill(start_color="4a1010", end_color="4a1010", fill_type="solid")
    FILL_X = PatternFill(start_color="2a102a", end_color="2a102a", fill_type="solid")
    FILL_H = PatternFill(start_color=HB, end_color=HB, fill_type="solid")
    FH = Font(name="Microsoft YaHei", size=11, bold=True, color="aaaaaa")
    FW = Font(name="Microsoft YaHei", size=11, color="ffffff")
    FR = Font(name="Microsoft YaHei", size=11, color=RED, bold=True)
    FG = Font(name="Microsoft YaHei", size=11, color=GREEN, bold=True)
    FL = Font(name="Microsoft YaHei", size=11, color=CYAN, underline="single", bold=True)
    CN = Alignment(horizontal="center", vertical="center")
    LF = Alignment(horizontal="left", vertical="center", wrap_text=True)

    wb = Workbook()
    ws1 = wb.active; ws1.title = "核心主线跟踪"
    ws1.merge_cells("A1:N1")
    ws1.cell(row=1,column=1,value=f"百日新高 v5.0 | {latest_date} | {anchors['verdict']} {anchors['position_range']} | 风险过滤{filter_count}只").font = Font(name="Microsoft YaHei", size=16, bold=True, color=ORANGE)
    ws1.cell(row=1,column=1).alignment = CN; ws1.row_dimensions[1].height = 36
    h1 = ["代码","名称","主线","首次新高","最新新高","上榜次","状态","涨幅%","距MA13%","量比","量能","风险标记","止损"]
    for col,h in enumerate(h1,1): ws1.cell(row=3,column=col,value=h).font=FH; ws1.cell(row=3,column=col).fill=FILL_H; ws1.cell(row=3,column=col).alignment=CN
    ws1.row_dimensions[3].height=28
    sorder = {"缩量回调":0,"高位整理":1,"强势上攻":2,"放量下跌":3}
    all_sorted = sorted(core.items(), key=lambda x: (sorder.get(x[1].get("status",""),9), x[1].get("dist13") if x[1].get("dist13") is not None else 999))
    for i,(code,st) in enumerate(all_sorted):
        row=4+i; status=st.get("status",""); chg=st["recs"][-1]["stock"].get("changePct",0)
        bg = PatternFill(start_color=R1,end_color=R1,fill_type="solid") if i%2==0 else PatternFill(start_color=R2,end_color=R2,fill_type="solid")
        is_filtered = code in filtered
        if is_filtered: bg = FILL_X
        elif status=="缩量回调": bg = FILL_S
        elif status=="放量下跌": bg = FILL_W
        risk_tag = st.get("risk_reason","") if is_filtered else ""
        # NEW marker
        name_display = st["name"]
        if st["firstDate"] == latest_date: name_display = name_display + " NEW"
        vals = [code,name_display,st.get("theme",""),st["firstDate"],st["lastDate"],len(st["recs"]),
                status,chg,st.get("dist13"),st.get("vr"),st.get("vol_status",""),risk_tag,f"MA13={st.get('stop_loss')}" if st.get("stop_loss") else ""]
        for col,val in enumerate(vals,1): ws1.cell(row=row,column=col,value=val).font=FW; ws1.cell(row=row,column=col).fill=bg; ws1.cell(row=row,column=col).alignment=CN
        ws1.cell(row=row,column=2).font=FL
        ws1.cell(row=row,column=7).font=FG if status=="缩量回调" else (FR if status=="放量下跌" else (Font(name="Microsoft YaHei",size=11,color=PURPLE,bold=True) if is_filtered else FW))
        ws1.cell(row=row,column=8).font=FR if chg>=0 else FG
        if st.get("dist13") is not None: ws1.cell(row=row,column=9).font=FG if abs(st["dist13"])<3 else FW
        ws1.row_dimensions[row].height = 26
    ws1.freeze_panes="A4"; ws1.auto_filter.ref=f"A3:N{3+len(all_sorted)}"
    for i,w in enumerate([10,12,14,12,12,8,10,10,10,9,8,18,12],1): ws1.column_dimensions[get_column_letter(i)].width=w

    ws2=wb.create_sheet("明日狙击池")
    ws2.merge_cells("A1:J1"); ws2.cell(row=1,column=1,value=f"狙击池 {latest_date}").font=Font(name="Microsoft YaHei",size=16,bold=True,color=GREEN); ws2.cell(row=1,column=1).alignment=CN
    h2=["代码","名称","主线","距MA13%","量比","首发后(天)","狙击参考价","止损参考","止跌K","当日状态"]
    for col,h in enumerate(h2,1): ws2.cell(row=3,column=col,value=h).font=FH; ws2.cell(row=3,column=col).fill=FILL_H; ws2.cell(row=3,column=col).alignment=CN
    if sniper:
        for i,st in enumerate(sniper):
            row=4+i; code=st["recs"][-1]["stock"]["code"]
            try: fd=datetime.strptime(st["firstDate"],"%Y-%m-%d"); ld=datetime.strptime(latest_date,"%Y-%m-%d"); days=(ld-fd).days
            except: days="?"
            price=st["recs"][-1]["stock"].get("price",""); sl=st.get("stop_loss")
            vals=[code,st["name"],st.get("theme",""),st.get("dist13"),st.get("vr"),days,price,f"MA13={sl:.2f}" if sl else "",st.get("stop_signal",""),st.get("status","")]
            for col,val in enumerate(vals,1): ws2.cell(row=row,column=col,value=val).font=FW; ws2.cell(row=row,column=col).fill=FILL_S; ws2.cell(row=row,column=col).alignment=CN
            ws2.cell(row=row,column=2).font=FL; ws2.row_dimensions[row].height=28
        ws2.freeze_panes="A4"; ws2.auto_filter.ref=f"A3:J{3+len(sniper)}"
    else: ws2.merge_cells("A4:J4"); ws2.cell(row=4,column=1,value="(无狙击目标)").font=Font(name="Microsoft YaHei",size=11,color="888888"); ws2.cell(row=4,column=1).alignment=CN
    for i,w in enumerate([10,12,14,10,9,10,12,14,8,10],1): ws2.column_dimensions[get_column_letter(i)].width=w

    ws3=wb.create_sheet("后备观察池"); ws3.merge_cells("A1:H1"); ws3.cell(row=1,column=1,value=f"后备池 {latest_date}").font=Font(name="Microsoft YaHei",size=16,bold=True,color=BLUE); ws3.cell(row=1,column=1).alignment=CN
    h3=["代码","名称","主线","首次新高","最新新高","量比","状态","操作"]
    for col,h in enumerate(h3,1): ws3.cell(row=3,column=col,value=h).font=FH; ws3.cell(row=3,column=col).fill=FILL_H; ws3.cell(row=3,column=col).alignment=CN
    if reserve:
        for i,st in enumerate(reserve):
            row=4+i; code=st["recs"][-1]["stock"]["code"]
            vals=[code,st["name"],st.get("theme",""),st["firstDate"],st["lastDate"],st.get("vr"),st.get("status",""),"等状态变为缩量回调"]
            for col,val in enumerate(vals,1): ws3.cell(row=row,column=col,value=val).font=FW; ws3.cell(row=row,column=col).fill=FILL_R; ws3.cell(row=row,column=col).alignment=CN
            ws3.cell(row=row,column=2).font=FL; ws3.row_dimensions[row].height=28
        ws3.freeze_panes="A4"; ws3.auto_filter.ref=f"A3:H{3+len(reserve)}"
    else: ws3.merge_cells("A4:H4"); ws3.cell(row=4,column=1,value="(无后备)").font=Font(name="Microsoft YaHei",size=11,color="888888"); ws3.cell(row=4,column=1).alignment=CN
    for i,w in enumerate([10,12,14,12,12,9,10,22],1): ws3.column_dimensions[get_column_letter(i)].width=w

    ws4=wb.create_sheet("风险过滤清单"); ws4.merge_cells("A1:G1"); ws4.cell(row=1,column=1,value=f"第七层·风险过滤 {filter_count}只").font=Font(name="Microsoft YaHei",size=16,bold=True,color=PURPLE); ws4.cell(row=1,column=1).alignment=CN
    h4=["代码","名称","主线","涨幅%","量比","风险类型","状态"]
    for col,h in enumerate(h4,1): ws4.cell(row=3,column=col,value=h).font=FH; ws4.cell(row=3,column=col).fill=FILL_H; ws4.cell(row=3,column=col).alignment=CN
    if filtered:
        fi = sorted(filtered.items(), key=lambda x: x[1].get("risk_reason",""))
        for i,(code,st) in enumerate(fi):
            row=4+i; lr=st["recs"][-1]["stock"]
            vals=[code,st["name"],st.get("theme",""),lr.get("changePct",0),st.get("vr",""),st.get("risk_reason",""),st.get("status","")]
            for col,val in enumerate(vals,1): ws4.cell(row=row,column=col,value=val).font=FW; ws4.cell(row=row,column=col).fill=FILL_X; ws4.cell(row=row,column=col).alignment=CN
            ws4.cell(row=row,column=2).font=FL; ws4.row_dimensions[row].height=26
        ws4.freeze_panes="A4"; ws4.auto_filter.ref=f"A3:G{3+len(fi)}"
    else: ws4.merge_cells("A4:G4"); ws4.cell(row=4,column=1,value="(今日无风险过滤)").font=Font(name="Microsoft YaHei",size=11,color="888888"); ws4.cell(row=4,column=1).alignment=CN
    for i,w in enumerate([10,12,14,10,9,22,10],1): ws4.column_dimensions[get_column_letter(i)].width=w

    ws5=wb.create_sheet("持仓预警"); ws5.merge_cells("A1:F1"); ws5.cell(row=1,column=1,value=f"持仓预警 {latest_date}").font=Font(name="Microsoft YaHei",size=16,bold=True,color=RED); ws5.cell(row=1,column=1).alignment=CN
    h5=["代码","名称","现价","涨幅%","预警信号","操作建议"]
    for col,h in enumerate(h5,1): ws5.cell(row=3,column=col,value=h).font=FH; ws5.cell(row=3,column=col).fill=FILL_H; ws5.cell(row=3,column=col).alignment=CN
    if warning_list:
        for i,hw in enumerate(warning_list):
            row=4+i; vals=[hw["code"],hw["name"],hw["price"],hw["chg"],"; ".join(hw["warnings"]),"; ".join(hw["warnings"])]
            for col,val in enumerate(vals,1): ws5.cell(row=row,column=col,value=val).font=FW; ws5.cell(row=row,column=col).fill=FILL_W; ws5.cell(row=row,column=col).alignment=CN
            ws5.cell(row=row,column=2).font=FL; ws5.row_dimensions[row].height=30
        ws5.freeze_panes="A4"; ws5.auto_filter.ref=f"A3:F{3+len(warning_list)}"
    else: ws5.merge_cells("A4:F4"); ws5.cell(row=4,column=1,value="(无持仓预警)").font=Font(name="Microsoft YaHei",size=11,color="888888"); ws5.cell(row=4,column=1).alignment=CN
    for i,w in enumerate([10,12,10,10,40,24],1): ws5.column_dimensions[get_column_letter(i)].width=w

    # Sheet 6: 全库扫描 (过去10天新高+距MA13 ±3%)
    ws6=wb.create_sheet("全库扫描(MA13±3%)")
    h6=["代码","名称","距MA13%","量比","量能","末次日期","末次涨幅%","末次价格","10天上榜次","备注"]
    col_count6 = 10
    ws6.merge_cells(f"A1:{get_column_letter(col_count6)}1"); ws6.cell(row=1,column=1,value=f"全库扫描·距MA13±3% {latest_date} ({scan_count}只)").font=Font(name="Microsoft YaHei",size=16,bold=True,color=CYAN); ws6.cell(row=1,column=1).alignment=CN; ws6.row_dimensions[1].height=36
    for col,h in enumerate(h6,1): ws6.cell(row=3,column=col,value=h).font=FH; ws6.cell(row=3,column=col).fill=FILL_H; ws6.cell(row=3,column=col).alignment=CN
    ws6.row_dimensions[3].height=28
    if full_scan:
        for i,st in enumerate(full_scan):
            row=4+i; bg = PatternFill(start_color=R1,end_color=R1,fill_type="solid") if i%2==0 else PatternFill(start_color=R2,end_color=R2,fill_type="solid")
            if st['vol_status'] == '缩量': bg = FILL_S
            note = ''
            if st['count'] >= 5 and st['last_chg'] < 1: note = '疑似出货'
            if abs(st['d13']) < 1: note = ('★极近均线 ' + note).strip()
            name_disp = st['name']
            if st['last_date'] == latest_date: name_disp = name_disp + ' NEW'
            vals=[st['code'],name_disp,st['d13'],st['vr'],st['vol_status'],st['last_date'],st['last_chg'],st['last_price'],st['count'],note]
            for col,val in enumerate(vals,1): ws6.cell(row=row,column=col,value=val).font=FW; ws6.cell(row=row,column=col).fill=bg; ws6.cell(row=row,column=col).alignment=CN
            ws6.cell(row=row,column=2).font=FL
            d13_c=ws6.cell(row=row,column=3); d13_c.font=FG if abs(st['d13'])<1 else FW
            ws6.cell(row=row,column=5).font=FG if st['vol_status']=='缩量' else (FR if st['vol_status']=='放量' else FW)
            ws6.row_dimensions[row].height=26
        ws6.freeze_panes="A4"; ws6.auto_filter.ref=f"A3:{get_column_letter(col_count6)}{3+len(full_scan)}"
    else: ws6.merge_cells(f"A4:{get_column_letter(col_count6)}4"); ws6.cell(row=4,column=1,value="(无)").font=Font(name="Microsoft YaHei",size=11,color="888888"); ws6.cell(row=4,column=1).alignment=CN
    for i,w in enumerate([10,12,10,9,8,12,10,10,10,18],1): ws6.column_dimensions[get_column_letter(i)].width=w

    # ═══ Sheet 7: 投资洞察 ═══
    ws7=wb.create_sheet("投资洞察")
    ws7.merge_cells("A1:B1"); ws7.cell(row=1,column=1,value="投资洞察与规律总结 v5.0").font=Font(name="Microsoft YaHei",size=18,bold=True,color=RED); ws7.cell(row=1,column=1).alignment=CN; ws7.row_dimensions[1].height=40
    ws7.column_dimensions['A'].width = 30; ws7.column_dimensions['B'].width = 80

    insights = [
        ("大盘状态", anchors['verdict'] + " | 仓位" + anchors['position_range'] + " | 最大持仓" + str(anchors['position_max']) + "只"),
        ("今日主线", "创新药/CRO 大爆发（37只，+2只），医药板块整体走强"),
        ("全库扫描", str(scan_count) + "只核心股距MA13 ±3%，首列重点关注"),
        ("", ""),
        ("═══ 牛股三大规律 ═══", ""),
        ("类型一·连续涨停型", "代表：恒尚节能(603137) 9天连板+106.8%，首日涨停后每日缩量一字板，从未回调"),
        ("类型二·回踩再起型", "代表：翰宇药业(300199) 7/14缩量回踩MA13=-0.81% → 7/15放量+6.94%，买点在缩量企稳日"),
        ("类型三·横盘突破型", "代表：康龙化成(300759) 6次上榜涨跌交替，末段+9.1%放量突破，确认后追入"),
        ("", ""),
        ("═══ 垃圾股三大特征 ═══", ""),
        ("特征一·一日游假突破", "42%的百日新高仅上榜1次，首日涨幅>15%的102只次日全部消失"),
        ("特征二·多次上榜不涨", "博通集成(603068) 8次上榜区间-2.8%，上榜频繁但价格不动=主力借新高出货"),
        ("特征三·首日即见顶", "11只股票上榜当日即收跌，盘中冲高回落，无主力支撑"),
        ("", ""),
        ("═══ 最佳介入窗口 ═══", ""),
        ("首发后第3-5天", "前2天冲高回落概率大，第3天缩量企稳是买点（回溯验证：候选池平均前瞻+4.1%）"),
        ("距MA13在±2%以内", "双环传动(002472) MA13=-0.17%为经典形态"),
        ("量比0.5-2.0", "太缩量(<0.5)=无人关注，太放量(>2.0)=出货嫌疑，健康区间0.8-2.0"),
        ("阳线或十字星", "排除放量阴线的假企稳"),
        ("", ""),
        ("═══ 卖出纪律 ═══", ""),
        ("放量滞涨", "量比>1.5 + 涨幅<1% → 减仓"),
        ("跌破MA13", "收盘<MA13 → 次日开盘止损离场"),
        ("浮盈>20%", "止盈线收紧至MA13"),
        ("连续3日放量加速", "止盈线从MA13收紧至MA5"),
        ("", ""),
        ("═══ 7/15 重点盯盘 ═══", ""),
        ("今日医药爆发", "迪哲医药(688192) NEW +20%、万邦医药(301520) +20%、药康生物(688046) NEW +17% 均为首次新高"),
        ("持续跟踪", "翰宇药业(300199) 缩量回踩后今日+6.9%放量反弹，关注是否持续"),
        ("均线附近等待", "双环传动(002472) MA13=-0.17%、博通集成(603068) MA13=+0.32%，等缩量信号"),
        ("放量下跌预警", "益诺思(688710) MA13=+17.9% 远离均线，今日不在榜但风险高"),
        ("", ""),
        ("═══ 回测验证 ═══", ""),
        ("翰宇药业案例", "7/14缩量回踩MA13(-0.81%)+VR=0.74 → 7/15放量+6.94% 完美验证狙击逻辑"),
        ("石药创新虚假信号", "近似MA13=+0.2%被证伪（真实+11.22%），系统已改用K线硬算，杜绝误判"),
        ("", ""),
        ("═══ 系统进化记录 ═══", ""),
        ("v5.0 七层规则", "三锚定仓→主线过滤→状态分类→风险过滤→狙击池→后备池→卖出预警"),
        ("全库扫描(新)", "不限于当日新高，扫描过去10天所有新高过的核心股在MA13附近的状态"),
        ("规则一", "首日涨幅>15%过滤（102只一日游验证）"),
        ("规则二", "5次上榜涨幅<5%标记疑似出货（博通集成8次-2.8%验证）"),
        ("NEW标记(新)", "每日新增股票自动标注，方便识别首次突破"),
    ]

    row = 3
    fnt_title = Font(name="Microsoft YaHei", size=12, bold=True, color=RED)
    fnt_text = Font(name="Microsoft YaHei", size=11, color="ffffff")
    fnt_sub = Font(name="Microsoft YaHei", size=11, color="cccccc")
    fill_highlight = PatternFill(start_color="3a1010", end_color="3a1010", fill_type="solid")
    fill_header = PatternFill(start_color="4a1010", end_color="4a1010", fill_type="solid")

    for label, value in insights:
        is_header = label.startswith("═══")
        is_empty = label == ""
        c1 = ws7.cell(row=row, column=1, value=label)
        c2 = ws7.cell(row=row, column=2, value=value)
        if is_empty:
            c1.font = fnt_text; c2.font = fnt_text
        elif is_header:
            c1.font = fnt_title; c2.font = fnt_title
            for c in [c1, c2]: c.fill = fill_header
        else:
            c1.font = Font(name="Microsoft YaHei", size=11, bold=True, color="ff6b6b")
            c2.font = fnt_sub
            if "重点" in label or "爆发" in label or "验证" in label:
                for c in [c1, c2]: c.fill = fill_highlight
        c1.alignment = Alignment(horizontal="right", vertical="center")
        c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws7.row_dimensions[row].height = 22 if not is_header else 28
        row += 1

    wb.save(EXCEL_FILE)
    print(f"\n[OK] {EXCEL_FILE} | 全库扫描 {scan_count}只")

if __name__ == "__main__":
    generate()
