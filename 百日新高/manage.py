#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
百日新高智能跟踪系统 v2.0
盯盘操作表：距买点距离 / 止损参考 / 量能判定 / 止跌检测 / 明日操作指令

用法：
  python manage.py report          → 生成 Excel
  python manage.py update-kline   → 用 wudao 批量拉取K线，计算MA/量比/止跌信号
  python manage.py show            → 终端查看汇总
"""
import json
import os
import sys
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.json")
EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "百日新高_报表.xlsx")
KLINE_CACHE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "kline_cache.json")

# ── 颜色体系 ──
DARK_BG = "0f1923"
HEADER_BG = "1a2a3a"
ROW1 = "1a2736"
ROW2 = "162230"
WHITE = "ffffff"
GRAY = "aaaaaa"
LIGHT_GRAY = "888888"
CYAN = "4fc3f7"

# 状态色
COLOR_STRONG = "ef5350"     # 强势上攻 - 红
COLOR_PULLBACK = "4caf50"   # 缩量回调 - 绿（机会）
COLOR_DIVERGE = "ffb74d"    # 分歧日 - 橙
COLOR_CONSOLIDATE = "81d4fa" # 高位整理 - 浅蓝
COLOR_BUY = "2196f3"        # 已触发买点 - 蓝
COLOR_HOLD = "ffd700"       # 已持仓 - 金
COLOR_DEFAULT = "cccccc"    # 默认

# 系统标记色
COLOR_STAR = "ffd700"       # ⭐ 机游共振
COLOR_DIAMOND = "4fc3f7"    # 💎 纯趋势
COLOR_FIRE = "ff6b35"       # 🔥 纯情绪

fill_status = {
    "强势上攻": PatternFill(start_color="4a1010", end_color="4a1010", fill_type="solid"),
    "缩量上涨": PatternFill(start_color="1a2a10", end_color="1a2a10", fill_type="solid"),
    "缩量回调": PatternFill(start_color="0a2a0a", end_color="0a2a0a", fill_type="solid"),
    "分歧日": PatternFill(start_color="3a2a10", end_color="3a2a10", fill_type="solid"),
    "高位整理": PatternFill(start_color="1a2a3a", end_color="1a2a3a", fill_type="solid"),
    "已触发买点": PatternFill(start_color="0a2a4a", end_color="0a2a4a", fill_type="solid"),
    "已持仓": PatternFill(start_color="3a3a00", end_color="3a3a00", fill_type="solid"),
}

font_status = {
    "强势上攻": Font(name="Microsoft YaHei", size=11, color=COLOR_STRONG, bold=True),
    "缩量上涨": Font(name="Microsoft YaHei", size=11, color="81c784", bold=True),
    "缩量回调": Font(name="Microsoft YaHei", size=11, color=COLOR_PULLBACK, bold=True),
    "分歧日": Font(name="Microsoft YaHei", size=11, color=COLOR_DIVERGE, bold=True),
    "高位整理": Font(name="Microsoft YaHei", size=11, color=COLOR_CONSOLIDATE, bold=True),
    "已触发买点": Font(name="Microsoft YaHei", size=11, color=COLOR_BUY, bold=True),
    "已持仓": Font(name="Microsoft YaHei", size=11, color=COLOR_HOLD, bold=True),
}

font_header = Font(name="Microsoft YaHei", size=11, bold=True, color=GRAY)
font_white = Font(name="Microsoft YaHei", size=11, color=WHITE)
font_red = Font(name="Microsoft YaHei", size=11, color=COLOR_STRONG, bold=True)
font_green = Font(name="Microsoft YaHei", size=11, color=COLOR_PULLBACK, bold=True)
font_link = Font(name="Microsoft YaHei", size=11, color=CYAN, underline="single", bold=True)
font_gray = Font(name="Microsoft YaHei", size=11, color=LIGHT_GRAY)
font_gray_sm = Font(name="Microsoft YaHei", size=9, color=LIGHT_GRAY)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center", wrap_text=True)


def load_data():
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def save_data(data):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_kline_cache():
    if os.path.exists(KLINE_CACHE):
        with open(KLINE_CACHE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_kline_cache(cache):
    with open(KLINE_CACHE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


def parse_main_net(val):
    if not val:
        return 0
    s = str(val).replace("+", "").replace(",", "")
    try:
        if "亿" in s:
            return float(s.replace("亿", "")) * 10000
        elif "万" in s:
            return float(s.replace("万", ""))
        return float(s) / 10000
    except:
        return 0


def parse_turnover(val):
    if not val:
        return 0
    s = str(val).replace("+", "").replace(",", "").replace("Z", "")
    try:
        if "亿" in s:
            return float(s.replace("亿", "")) * 10000
        elif "万" in s:
            return float(s.replace("万", ""))
        return float(s) / 10000
    except:
        return 0


# ═══════════════════════════
# MA / 量比 / 止跌K线计算
# ═══════════════════════════

def calc_ma(closes, period):
    """计算移动均线"""
    if len(closes) < period:
        return None
    return sum(closes[-period:]) / period


def detect_stop_kline(row):
    """检测止跌K线信号：十字星/小阳线/长下影线"""
    if not row or len(row) < 4:
        return False
    o, c, h, l = row[0], row[1], row[2], row[3]
    body = abs(c - o)
    upper_shadow = h - max(o, c)
    lower_shadow = min(o, c) - l
    total_range = h - l
    if total_range == 0:
        return False

    body_ratio = body / total_range
    lower_ratio = lower_shadow / total_range if total_range > 0 else 0

    # 十字星：实体占比 < 20%
    if body_ratio < 0.2:
        return True
    # 长下影线：下影线 > 实体 * 2 且下影线占比 > 40%
    if lower_ratio > 0.4 and lower_shadow > body * 2:
        return True
    # 小阳线止跌：阳线实体 20%-50% 且收涨
    if c > o and 0.2 <= body_ratio <= 0.5:
        return True
    return False


def compute_indicators(kline_rows):
    """从K线数据计算技术指标
    返回: {distMA13, distMA20, ma13_price, ma20_price, volRatio, volStatus, stopSignal, stopLoss}
    """
    if not kline_rows or len(kline_rows) < 20:
        return None

    closes = [r[1] for r in kline_rows]  # 收盘价
    volumes = [r[4] if len(r) > 4 else 0 for r in kline_rows]  # 成交量
    latest_close = closes[-1]
    latest_vol = volumes[-1] if volumes else 0

    ma13 = calc_ma(closes, 13)
    ma20 = calc_ma(closes, 20)
    avg_vol_5 = sum(volumes[-6:-1]) / 5 if len(volumes) >= 6 else (sum(volumes[:-1]) / max(1, len(volumes) - 1))

    dist13 = round((latest_close / ma13 - 1) * 100, 2) if ma13 and ma13 > 0 else None
    dist20 = round((latest_close / ma20 - 1) * 100, 2) if ma20 and ma20 > 0 else None
    vol_ratio = round(latest_vol / avg_vol_5, 2) if avg_vol_5 > 0 else None

    if vol_ratio is not None:
        if vol_ratio < 0.8:
            vol_status = "缩量"
        elif vol_ratio > 1.2:
            vol_status = "放量"
        else:
            vol_status = "正常"
    else:
        vol_status = ""

    # 止跌K线检测（最近一根K线）
    latest_kline = kline_rows[-1]
    stop_signal = detect_stop_kline(latest_kline)

    # 止损参考价 = MA13 或 买入日最低价 -3%（取两者较高者）
    stop_loss = round(ma13, 2) if ma13 else None

    return {
        "distMA13": dist13,
        "distMA20": dist20,
        "ma13_price": round(ma13, 2) if ma13 else None,
        "ma20_price": round(ma20, 2) if ma20 else None,
        "volRatio": vol_ratio,
        "volStatus": vol_status,
        "stopSignal": stop_signal,
        "stopLoss": stop_loss,
    }


# ═══════════════════════════
# 状态判定
# ═══════════════════════════

def determine_status(stock, indicators, is_holding=False):
    """根据涨幅+量比判定股票状态"""
    if is_holding:
        return "已持仓"

    chg = stock.get("changePct", 0)
    vol_ratio = indicators.get("volRatio") if indicators else None

    # 强势上攻：涨幅 >= 5% 且放量
    if chg >= 5:
        if vol_ratio and vol_ratio > 1.5:
            return "强势上攻"
        return "强势上攻"

    # 缩量回调：涨幅 < 0 且缩量
    if chg < 0:
        if vol_ratio and vol_ratio < 0.8:
            return "缩量回调"
        elif vol_ratio and vol_ratio > 1.5:
            return "分歧日"
        return "缩量回调"

    # 缩量上涨：涨幅 >= 0 但不大，缩量
    if chg >= 0 and chg < 3:
        if vol_ratio and vol_ratio < 0.8:
            return "缩量上涨"
        return "高位整理"

    # 其他
    if chg >= 3:
        return "强势上攻"

    return "高位整理"


# ═══════════════════════════
# 明日操作建议生成
# ═══════════════════════════

def generate_action(status, indicators, stock):
    """根据状态和指标生成操作建议"""
    if status == "已持仓":
        return "持有，止损参考MA13"

    dist13 = indicators.get("distMA13") if indicators else None
    dist20 = indicators.get("distMA20") if indicators else None
    stop_sig = indicators.get("stopSignal", False) if indicators else False
    vol_ratio = indicators.get("volRatio") if indicators else None
    chg = stock.get("changePct", 0)

    if status == "缩量回调":
        if stop_sig and dist13 is not None and dist13 < 3:
            return "狙击目标：次日开盘可低吸"
        elif dist13 is not None and abs(dist13) < 2:
            addr = "已靠近MA13" if dist13 >= 0 else f"已跌破MA13 {abs(dist13):.1f}%"
            return f"等缩量回踩MA13企稳，出现十字星可低吸（{addr}）"
        elif dist13 is not None:
            return f"继续观察，等回踩MA13（距MA13 {dist13:+.1f}%）"
        else:
            return "缩量回调中，需补充K线数据计算MA距离"

    if status == "强势上攻":
        return "不宜追高，等缩量回踩"

    if status == "缩量上涨":
        return "观望，等放量突破或缩量回踩"

    if status == "分歧日":
        return "分歧中，观望为宜"

    if status == "高位整理":
        return "横盘整理，方向不明"

    return "待观察"


# ═══════════════════════════
# Excel 生成
# ═══════════════════════════

def write_title_row(ws, row, text, cols=18):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Microsoft YaHei", size=16, bold=True, color="ff6b35")
    c.alignment = center
    ws.row_dimensions[row].height = 36


def write_headers(ws, row, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = font_header
        c.fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
        c.alignment = center
        c.border = Border(bottom=Side(style="thin", color="444444"))
    ws.row_dimensions[row].height = 28


def generate_excel():
    data = load_data()
    kline_cache = load_kline_cache()
    records = data["records"]

    # ── 构建股票持仓状态 ──
    holdings = set()
    if "holdings" in data:
        holdings = set(data["holdings"])

    # ── 汇总所有股票 ──
    stock_info = {}  # code -> {name, firstDate, lastDate, dates, sectors, records:[{date,stock}]}
    for rec in records:
        date = rec["date"]
        for s in rec["stocks"]:
            code = s["code"]
            if code not in stock_info:
                stock_info[code] = {
                    "name": s["name"],
                    "firstDate": date,
                    "lastDate": date,
                    "dates": [],
                    "sectors": set(),
                    "records": [],
                    "mainNet_total": 0,
                }
            info = stock_info[code]
            info["dates"].append(date)
            info["lastDate"] = date
            info["sectors"].add(s.get("sector", ""))
            info["records"].append({"date": date, "stock": s})
            info["mainNet_total"] += parse_main_net(s.get("mainNet", ""))

    # 排序
    for code, info in stock_info.items():
        info["dates"].sort()

    # ── 按回调→买点 优先排序 ──
    def sort_key(item):
        code, info = item
        latest_rec = info["records"][-1]["stock"]
        chg = latest_rec.get("changePct", 0)

        cached = kline_cache.get(code, {})
        indicators = cached.get("indicators", None) if cached else None
        dist = indicators.get("distMA13", 99) if indicators else 99

        status = determine_status(latest_rec, indicators)
        score = 0
        if status == "缩量回调":
            score = 100 - abs(dist or 99) * 5  # 离均线越近越优先
        elif status == "缩量上涨":
            score = 60 - abs(dist or 99) * 3
        elif status == "分歧日":
            score = 40
        elif status == "高位整理":
            score = 20
        score += info["records"].__len__() * 5  # 多次上榜加分
        return -score

    sorted_stocks = sorted(stock_info.items(), key=sort_key)

    wb = Workbook()

    # ═══════════════ Sheet 1: 跟踪总表 ═══════════════
    ws1 = wb.active
    ws1.title = "跟踪总表"

    write_title_row(ws1, 1, "百日新高 · 盯盘操作表")

    # 统计行
    total = len(stock_info)
    pullback_count = sum(1 for code, info in stock_info.items()
                         if determine_status(info["records"][-1]["stock"],
                                             kline_cache.get(code, {}).get("indicators"))
                         in ["缩量回调", "缩量上涨"])
    held = len(holdings)
    ws1.merge_cells("A1:R1")
    ws1.merge_cells("A2:R2")
    stat_text = f"总{total}只 | 缩量回踩候选{pullback_count}只 | 已持仓{held}只 | 覆盖{len(records)}个交易日 | {records[-1]['date']}收盘"
    ws1.cell(row=2, column=1, value=stat_text).font = font_gray
    ws1.cell(row=2, column=1).alignment = center
    ws1.row_dimensions[2].height = 22

    headers1 = [
        "代码", "名称", "所属主线", "系统标记",
        "首次新高日", "最新新高日", "上榜次数",
        "当前状态", "今日涨幅%",
        "距MA13%", "距MA20%", "今日量比", "量能判定",
        "止跌K线", "止损参考价",
        "明日操作",
        "主力净额", "备注",
    ]
    write_headers(ws1, 4, headers1)
    cols_count = len(headers1)

    for i, (code, info) in enumerate(sorted_stocks):
        row = 5 + i
        bg = fill_status.get("高位整理", PatternFill(start_color=ROW1, end_color=ROW1, fill_type="solid"))
        bg = PatternFill(start_color=ROW1, end_color=ROW1, fill_type="solid") if i % 2 == 0 else PatternFill(start_color=ROW2, end_color=ROW2, fill_type="solid")

        latest_rec = info["records"][-1]["stock"]
        chg = latest_rec.get("changePct", 0)

        # K线指标
        cached = kline_cache.get(code, {})
        indicators = cached.get("indicators", None) if cached else None

        # 状态判定
        is_holding = code in holdings
        status = determine_status(latest_rec, indicators, is_holding)

        # 系统标记（用户可设置，默认按板块推断）
        sys_mark = cached.get("sysMark", "") if cached else ""
        main_line = cached.get("mainLine", latest_rec.get("sector", ""))

        # 量能
        vol_ratio = indicators.get("volRatio") if indicators else None
        vol_status = indicators.get("volStatus", "") if indicators else ""
        dist13 = indicators.get("distMA13") if indicators else None
        dist20 = indicators.get("distMA20") if indicators else None
        stop_sig = "是" if (indicators and indicators.get("stopSignal")) else ""
        stop_loss = indicators.get("stopLoss") if indicators else None
        stop_loss_str = f"MA13={stop_loss:.2f}" if stop_loss else ""

        # 明日操作
        action = generate_action(status, indicators, latest_rec)
        if is_holding and not action:
            action = "持有中"

        values = [
            code,
            info["name"],
            main_line,
            sys_mark,
            info["firstDate"],
            info["lastDate"],
            info["records"].__len__(),
            status,
            chg,
            dist13,
            dist20,
            vol_ratio,
            vol_status,
            stop_sig,
            stop_loss_str,
            action,
            latest_rec.get("mainNet", ""),
            cached.get("notes", ""),
        ]

        for col, val in enumerate(values, 1):
            c = ws1.cell(row=row, column=col, value=val)
            c.font = font_white
            c.alignment = center

        # 整行状态底色
        status_fill = fill_status.get(status, bg)
        for col in range(1, cols_count + 1):
            ws1.cell(row=row, column=col).fill = status_fill

        # 代码列
        ws1.cell(row=row, column=1).font = font_white

        # 名称链接
        ws1.cell(row=row, column=2).font = font_link

        # 状态列
        status_cell = ws1.cell(row=row, column=8)
        status_cell.font = font_status.get(status, font_white)

        # 涨幅颜色
        chg_c = ws1.cell(row=row, column=9)
        chg_c.font = font_red if chg >= 0 else font_green

        # MA距离颜色（距均线越近越安全）
        for col, dist_val in [(10, dist13), (11, dist20)]:
            dist_cell = ws1.cell(row=row, column=col)
            if dist_val is not None:
                if abs(dist_val) < 2:
                    dist_cell.font = Font(name="Microsoft YaHei", size=11, color=COLOR_PULLBACK, bold=True)
                elif abs(dist_val) < 5:
                    dist_cell.font = Font(name="Microsoft YaHei", size=11, color=COLOR_DIVERGE)
                else:
                    dist_cell.font = font_white

        # 量比颜色
        vol_c = ws1.cell(row=row, column=12)
        if vol_ratio is not None:
            if vol_ratio < 0.8:
                vol_c.font = font_green
            elif vol_ratio > 1.5:
                vol_c.font = font_red

        # 止跌K线
        stop_c = ws1.cell(row=row, column=14)
        if stop_sig == "是":
            stop_c.font = Font(name="Microsoft YaHei", size=11, color=COLOR_PULLBACK, bold=True)

        # 止损参考
        sl_c = ws1.cell(row=row, column=15)
        if stop_loss:
            sl_c.font = Font(name="Microsoft YaHei", size=10, color="ffaa00")

        # 明日操作列 - 加粗
        act_c = ws1.cell(row=row, column=16)
        act_c.font = Font(name="Microsoft YaHei", size=10, color="ffd700" if "狙击" in str(action) else LIGHT_GRAY)
        act_c.alignment = left

        # 备注列
        ws1.cell(row=row, column=18).alignment = left

        ws1.row_dimensions[row].height = 28

    ws1.freeze_panes = "E5"
    last_data_row1 = 4 + len(sorted_stocks)
    ws1.auto_filter.ref = f"A4:R{last_data_row1}"

    col_widths_1 = {
        1: 10, 2: 12, 3: 16, 4: 10, 5: 12, 6: 12, 7: 8,
        8: 12, 9: 10, 10: 10, 11: 10, 12: 9, 13: 8,
        14: 8, 15: 14, 16: 32, 17: 12, 18: 16,
    }
    for col, w in col_widths_1.items():
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ═══════════════ Sheet 2: 缩量回调候选 ═══════════════
    ws2 = wb.create_sheet("缩量回调候选")
    write_title_row(ws2, 1, "缩量回调 · 狙击候选")

    pullback_stocks = []
    for code, info in sorted_stocks:
        latest_rec = info["records"][-1]["stock"]
        cached = kline_cache.get(code, {})
        indicators = cached.get("indicators", None) if cached else None
        status = determine_status(latest_rec, indicators)
        if status in ["缩量回调", "缩量上涨"]:
            pullback_stocks.append((code, info, status, indicators))

    pullback_stocks.sort(key=lambda x: (x[3].get("distMA13", 99) if x[3] else 99))

    ws2.merge_cells("A2:M2")
    ws2.cell(row=2, column=1, value=f"候选{pullback_stocks.__len__()}只 | 按距MA13由近到远排列").font = font_gray
    ws2.cell(row=2, column=1).alignment = center

    headers2 = ["代码", "名称", "所属主线", "状态", "今日涨幅%", "距MA13%", "距MA20%",
                "今日量比", "量能判定", "止跌K线", "止损参考价", "明日操作", "备注"]
    write_headers(ws2, 4, headers2)

    for i, (code, info, status, indicators) in enumerate(pullback_stocks):
        row = 5 + i
        bg = PatternFill(start_color=ROW1, end_color=ROW1, fill_type="solid") if i % 2 == 0 else PatternFill(start_color=ROW2, end_color=ROW2, fill_type="solid")
        latest_rec = info["records"][-1]["stock"]
        chg = latest_rec.get("changePct", 0)

        cached = kline_cache.get(code, {})
        dist13 = indicators.get("distMA13") if indicators else None
        dist20 = indicators.get("distMA20") if indicators else None
        vol_ratio = indicators.get("volRatio") if indicators else None
        vol_status = indicators.get("volStatus", "") if indicators else ""
        stop_sig = "是" if (indicators and indicators.get("stopSignal")) else ""
        stop_loss = indicators.get("stopLoss") if indicators else None
        action = generate_action(status, indicators, latest_rec)

        values = [code, info["name"], cached.get("mainLine", latest_rec.get("sector", "")),
                  status, chg, dist13, dist20, vol_ratio, vol_status, stop_sig,
                  f"MA13={stop_loss:.2f}" if stop_loss else "", action, cached.get("notes", "")]
        for col, val in enumerate(values, 1):
            c = ws2.cell(row=row, column=col, value=val)
            c.font = font_white
            c.fill = bg
            c.alignment = center

        ws2.cell(row=row, column=2).font = font_link
        ws2.cell(row=row, column=4).font = font_status.get(status, font_white)
        ws2.cell(row=row, column=5).font = font_red if chg >= 0 else font_green
        ws2.cell(row=row, column=12).font = Font(name="Microsoft YaHei", size=10, color=COLOR_PULLBACK)

        if dist13 is not None and abs(dist13) < 3:
            for col2 in range(1, 14):
                ws2.cell(row=row, column=col2).fill = PatternFill(start_color="0a1a0a", end_color="0a1a0a", fill_type="solid")

        ws2.row_dimensions[row].height = 28

    ws2.freeze_panes = "A5"
    ws2.auto_filter.ref = f"A4:M{4 + max(1, pullback_stocks.__len__())}"
    widths2 = [10, 12, 16, 10, 10, 10, 10, 9, 8, 8, 14, 32, 16]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ═══════════════ Sheet 3: 板块热度 ═══════════════
    ws3 = wb.create_sheet("板块热度")
    write_title_row(ws3, 1, "板块热度分析", cols=12)

    sector_by_date = defaultdict(lambda: defaultdict(int))
    all_sectors = set()
    for rec in records:
        for s in rec["stocks"]:
            for sec in s.get("sector", "").split():
                sector_by_date[rec["date"]][sec] += 1
                all_sectors.add(sec)

    dates = sorted(set(r["date"] for r in records))
    sector_total = {}
    for sec in all_sectors:
        total = sum(sector_by_date[d].get(sec, 0) for d in dates)
        if total >= 2:
            sector_total[sec] = total
    sorted_sectors = sorted(sector_total.items(), key=lambda x: -x[1])

    headers3 = ["板块", "累计次数"] + dates
    write_headers(ws3, 3, headers3)

    for i, (sec, total) in enumerate(sorted_sectors):
        row = 4 + i
        bg = PatternFill(start_color=ROW1, end_color=ROW1, fill_type="solid") if i % 2 == 0 else PatternFill(start_color=ROW2, end_color=ROW2, fill_type="solid")
        values = [sec, total] + [sector_by_date[d].get(sec, 0) for d in dates]
        for col, val in enumerate(values, 1):
            c = ws3.cell(row=row, column=col, value=val)
            c.font = font_white
            c.fill = bg
            c.alignment = center
        ws3.row_dimensions[row].height = 24

    ws3.freeze_panes = "C4"
    ws3.auto_filter.ref = f"A3:{get_column_letter(len(headers3))}{3 + len(sorted_sectors)}"
    for i in range(1, len(headers3) + 1):
        ws3.column_dimensions[get_column_letter(i)].width = 14

    # ═══════════════ Sheet 4: 已持仓 ═══════════════
    ws4 = wb.create_sheet("已持仓")
    write_title_row(ws4, 1, "已持仓管理")

    if holdings:
        held_stocks = [(code, info) for code, info in stock_info.items() if code in holdings]
        headers4 = ["代码", "名称", "持仓日期", "持仓价", "现价", "盈亏%", "止损价", "所属主线", "备注"]
        write_headers(ws4, 3, headers4)
        for i, (code, info) in enumerate(held_stocks):
            row = 4 + i
            bg = PatternFill(start_color=ROW1, end_color=ROW1, fill_type="solid")
            values = [code, info["name"], "", "", info["records"][-1]["stock"]["price"], "", "",
                      ";".join(info["sectors"]), ""]
            for col, val in enumerate(values, 1):
                c = ws4.cell(row=row, column=col, value=val)
                c.font = font_white
                c.fill = bg
                c.alignment = center
            ws4.cell(row=row, column=2).font = font_link
            ws4.row_dimensions[row].height = 28
        ws4.freeze_panes = "A4"
        ws4.auto_filter.ref = f"A3:I{3 + len(held_stocks)}"
    else:
        ws4.merge_cells("A3:H3")
        ws4.cell(row=3, column=1, value="(暂无持仓)").font = font_gray
        ws4.cell(row=3, column=1).alignment = center

    widths4 = [10, 12, 12, 10, 10, 10, 10, 20, 16]
    for i, w in enumerate(widths4, 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    # ═══════════════ Sheet 5: 每日明细 ═══════════════
    ws5 = wb.create_sheet("每日明细")
    write_title_row(ws5, 1, "每日原始数据")
    headers5 = ["日期", "股票名称", "代码", "价格", "涨幅%", "板块", "成交额", "主力净额"]
    write_headers(ws5, 3, headers5)

    row5 = 4
    for rec in records:
        for stock in rec["stocks"]:
            bg = PatternFill(start_color=ROW1, end_color=ROW1, fill_type="solid") if row5 % 2 == 0 else PatternFill(start_color=ROW2, end_color=ROW2, fill_type="solid")
            values = [rec["date"], stock["name"], stock["code"], stock["price"], stock["changePct"],
                      stock["sector"], stock["turnover"], stock["mainNet"]]
            for col, val in enumerate(values, 1):
                c = ws5.cell(row=row5, column=col, value=val)
                c.font = font_white
                c.fill = bg
                c.alignment = center
            ws5.cell(row=row5, column=2).font = font_link
            chg_c = ws5.cell(row=row5, column=5)
            chg_c.font = font_red if stock.get("changePct", 0) >= 0 else font_green
            ws5.row_dimensions[row5].height = 24
            row5 += 1
        row5 += 1

    ws5.freeze_panes = "A4"
    ws5.auto_filter.ref = f"A3:H{row5 - 1}"
    widths5 = [14, 14, 10, 10, 10, 22, 14, 14]
    for i, w in enumerate(widths5, 1):
        ws5.column_dimensions[get_column_letter(i)].width = w

    wb.save(EXCEL_FILE)
    print(f"[OK] Excel: {EXCEL_FILE}")
    print(f"   跟踪总表: {total}只 | 缩量回调候选 {pullback_count}只 | 已持仓 {held}只")


def cmd_report():
    generate_excel()


def cmd_show():
    data = load_data()
    kline_cache = load_kline_cache()
    stock_info = {}
    for rec in data["records"]:
        for s in rec["stocks"]:
            code = s["code"]
            if code not in stock_info:
                stock_info[code] = {"name": s["name"], "count": 0, "dates": [], "sectors": set(), "latest": s}
            stock_info[code]["count"] += 1
            stock_info[code]["dates"].append(rec["date"])
            stock_info[code]["sectors"].add(s.get("sector", ""))
            stock_info[code]["latest"] = s

    print("\n=== 百日新高 v2.0 汇总 ===")
    print(f"交易日: {len(data['records'])}天")
    print(f"总个股: {len(stock_info)}只")

    # 缩量回调
    pullback = []
    for code, info in stock_info.items():
        cached = kline_cache.get(code, {})
        indicators = cached.get("indicators", None) if cached else None
        status = determine_status(info["latest"], indicators)
        if status in ["缩量回调", "缩量上涨"]:
            pullback.append((code, info, indicators, status))
    pullback.sort(key=lambda x: (x[2].get("distMA13", 99) if x[2] else 99))

    if pullback:
        print(f"\n--- 缩量回调候选 (按距MA13距离) ---")
        for code, info, indicators, status in pullback:
            dist = indicators.get("distMA13", "?") if indicators else "?"
            stop = f", 止跌K线" if (indicators and indicators.get("stopSignal")) else ""
            print(f"  [{status}] {info['name']}({code}) 距MA13:{dist}%{stop}")

    # 连续上榜
    multi = [(code, info) for code, info in stock_info.items() if info["count"] >= 2]
    if multi:
        multi.sort(key=lambda x: -x[1]["count"])
        print(f"\n--- 多次上榜 ---")
        for code, info in multi[:10]:
            print(f"  [{info['count']}次] {info['name']}({code})  {';'.join(list(info['sectors'])[:2])}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
    elif sys.argv[1] == "report":
        cmd_report()
    elif sys.argv[1] == "show":
        cmd_show()
    else:
        print(f"Unknown command: {sys.argv[1]}")
